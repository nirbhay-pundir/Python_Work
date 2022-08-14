import pyautogui as py
import time
import openpyxl


def click(x, y):
    py.moveTo(x, y, 0.2)
    py.click(x, y)


def sendMail():
    # Compose Button
    click(595, 897)

    # Move to subject text box and click
    py.moveTo(318, 248, 1)
    py.click(318, 248)

    # Open clipboard and paste email
    click(536, 716)
    click(362, 815)
    click(327, 324)

    # Move to subject and click
    click(327, 324)
    click(619, 714)

    # Paste subject from clipboard
    click(478, 794)
    click(570, 810)

    # Move and click on email box
    click(370, 400)

    # Paste email from clipboard
    click(619, 714)
    click(478, 794)
    click(560, 940)

    # Send
    click(620, 100)


wb = openpyxl.load_workbook("U1LoginEmails.xlsx")
sheet = wb['Sheet1']
x = 313
y = 426
# y = 972
Y = 222 + 25
# Y = 1028
Y1 = 222 + 25
# Y1 = 1028
Y2 = 222 + 25
# Y2 = 1028
Y3 = 222 + 25
# Y3 = 1028
Y4 = 480 + 25

drag = 150
py.click(150, 16)

print(1, ": ", sheet.cell(1, 1).value)
for i in range(2, 56):
    sendMail()
    print(i, ": ", sheet.cell(i, 1).value)

    click(650, 104)

    for t in reversed(range(0, 10)):
        print("Waiting for " + str(t) + " sec\r", end="")
        time.sleep(1)

    if y > 972:
        if Y > 950:
            Y = 951
            py.moveTo(x, 950, 0.5)
            py.dragTo(x, 110, 1.5)
            py.moveTo(x, 950, 0.5)
            py.dragTo(x, 110, 1.5)
            if Y1 > 950:
                Y1 = 951
                py.moveTo(x, 950, 0.5)
                py.dragTo(x, 110, 1.5)
                if Y2 > 950:
                    Y2 = 951
                    py.moveTo(x, 950, 0.5)
                    py.dragTo(x, 110, 1.5)
                    if Y3 > 950:
                        Y3 = 951
                        py.moveTo(x, 950, 0.5)
                        py.dragTo(x, 110, 1.5)
                        click(x, Y4)
                        Y4 += 78
                    else:
                        py.moveTo(x, Y3, 0.5)
                        py.click(x, Y3)
                        Y3 += 78
                else:
                    click(x, Y2)
                    Y2 += 78
            else:
                click(x, Y1)
                Y1 += 78
        else:
            y = 973
            py.moveTo(x, y, 0.5)
            py.dragTo(x, 120, 1.5)
            click(x, Y)
            Y += 78
    else:
        click(x, y)
        y += 78
