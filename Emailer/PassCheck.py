import openpyxl
import smtplib
import time
from email.message import EmailMessage

wb = openpyxl.load_workbook("list.xlsx")
sheet = wb['Sheet1']

def passwordcheck(usr,psw):
    server = smtplib.SMTP('smtp.gmail.com',587)
    server.starttls()
    try:
        server.login(usr,psw)
        email = EmailMessage()
        email['From'] = 'Sender_Email'
        email['To'] = "gmail.com"
        email['Subject'] = "subject"
        email.set_content("message")
        server.send_message(email)
        print("Send : ",end="")
        ret = True
    except:
        ret = False
    server.quit()
    return ret

for i in range(1,2):
    print(i,": ",sheet.cell(i,1).value," : ",sheet.cell(i,2).value," : ",end="")
    time.sleep(1)
    if passwordcheck(sheet.cell(i,1).value,sheet.cell(i,2).value):
        sheet.cell(i, 3, value = "Verified")
        print("Correct")
    else:
        print("Incorrect")
    wb.save('result.xlsx')
    for t in reversed(range(0, 5)):
        print("Waiting for " + str(t) + " sec\r", end="")
        time.sleep(1)