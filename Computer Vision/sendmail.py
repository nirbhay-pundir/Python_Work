import random
import time
import openpyxl

import cv2
import mss
import pyautogui as py
import numpy as np
import pytesseract


def matchImg(template):
    time.sleep(1)
    template = cv2.cvtColor(template, cv2.COLOR_RGB2GRAY)
    img = np.array(py.screenshot())
    img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    threshold = 0.9
    loc = np.where(res >= threshold)
    return loc[0][0] + 25, loc[1][0] + 25


def click(x, y):
    py.moveTo(x, y, 0.5)
    py.click(x, y)


def getCoordinates(i):
    while True:
        img = np.array(sct.grab(mon))
        data = pytesseract.image_to_data(img)
        k = 0
        for a, b in enumerate(data.splitlines()):
            if a != 0:
                b = b.split()
                if len(b) == 12:
                    # print(b)
                    if b[11] == sheet.cell(i, 1).value:
                        # print(b)
                        # print(b[11], " : Clicked")
                        return int(b[6]), int(b[7])
                    elif k == 6:
                        py.scroll(-100)
                    k += 1


def sendMail():
    # Move to sender text box and click
    py.moveTo(318, 248, 0.5)
    py.click(318, 248)

    # Open clipboard and paste email
    click(536, 716)
    # click(327, 324)
    # Paste all mails
    click(346, 877)
    # Paste yhtk01
    click(536, 716)
    click(362, 815)
    click(273, 714)
    click(677, 973)
    # time.sleep(0.2)

    # Click Enter
    click(677, 973)

    # Move to subject and click
    # click(327, 324)
    click(300, 387)
    click(619, 714)

    # Paste subject from clipboard
    click(478, 794)
    click(570, 810)

    # Move and click on email box
    click(490, 490)

    # Paste email from clipboard
    click(619, 714)
    click(478, 794)
    click(560, 940)

    # Send
    # click(620, 100)

    # Schedule send
    click(683, 103)
    click(537, 117)
    # click(574, 576)
    l, m = matchImg(cv2.imread('icon.png'))
    click(m, l)
    click(375, 570)
    click(340, 768)
    # Set time hour
    click(341, 590)
    click(653, 915)
    click(653, 915)
    n = random.randint(10, 13)
    # n=12
    if n == 13:
        n = 1
        py.typewrite(str(n))
        click(550,445)
        click(550,506)
    elif n==12:
        py.typewrite(str(n))
        click(550, 445)
        click(550, 506)
    else:
        py.typewrite(str(n))
    # Set time minutes
    click(430, 440)
    click(653, 915)
    click(653, 915)
    click(653, 915)
    py.typewrite(str(random.randint(0, 59)))

    click(598, 530)
    click(610, 690)
    time.sleep(2)


pytesseract.pytesseract.tesseract_cmd = "tesseract"
wb = openpyxl.load_workbook("U1LoginEmails.xlsx")
sheet = wb['Sheet1']
sct = mss.mss()
# # select window
click(150, 16)
# # Compose Button
# click(595, 910)
# click(280,100)
# print(1, ": ", sheet.cell(1, 1).value,end=": ")
# sendMail()
# print("Done!")

# Capture screen Coordinates
left = 340
top = 68
right = 712
bottom = 680
mon = {"top": top, "left": left, "width": right - left, "height": bottom - top}

# py.moveTo(400, 600, 0.5)

for i in range(2, 56):
    print(i, ": ", sheet.cell(i, 1).value, end=": ")
    l, m = matchImg(cv2.imread('compose.png'))
    click(m, l)
    # click(595, 910)
    click(640, 180)
    # print(i)
    x, y = getCoordinates(i)
    # print(x+left, " ", y+top)
    click(x + left + 50, y + top + 30)
    # click(280,100)
    # time.sleep(1)
    # click(595, 910)
    sendMail()
    print("Done!")
